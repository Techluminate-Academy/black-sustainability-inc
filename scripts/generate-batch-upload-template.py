#!/usr/bin/env python3
"""
Generate BSN Batch Upload Template with Dropdown Validation
This script creates an Excel file with data validation dropdowns for:
- IDENTIFICATION
- GENDER  
- PRIMARY INDUSTRY HOUSE
- ADDITIONAL FOCUS AREAS
- State/Province
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

def create_batch_upload_template():
    """Create Excel template with dropdown validations"""
    
    # Create a new workbook
    wb = Workbook()
    
    # Main data sheet
    ws_data = wb.active
    ws_data.title = "Batch Upload Template"
    
    # Create options sheet (hidden)
    ws_options = wb.create_sheet("Options")
    
    # Define column headers
    headers = [
        "EMAIL ADDRESS",
        "Email 2", 
        "FIRST NAME",
        "LAST NAME",
        "ORGANIZATION NAME",
        "WEBSITE",
        "BIO",
        "IDENTIFICATION",
        "GENDER",
        "PHONE US/CAN ONLY",
        "PHONE NON-US/CAN",
        "PRIMARY INDUSTRY HOUSE",
        "ADDITIONAL FOCUS AREAS",
        "NAICS Code",
        "AFFILIATED ENTITY",
        "Address",
        "Location (Nearest City)",
        "Country",
        "State/Province",
        "State",
        "Zip/Postal Code",
        "Time zone",
        "Include me on Global BSN Map",
        "LATITUDE (NEW)",
        "LONGITUDE (NEW)",
        "MEMBER LEVEL",
        "Paying Member (keep current)",
        "Equity Member (keep current)",
        "Membership Status Notes",
        "Send Need Payment Email"
    ]
    
    # Write headers with formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Adjust column widths
    column_widths = {
        'A': 30, 'B': 25, 'C': 15, 'D': 15, 'E': 30, 'F': 30, 'G': 50,
        'H': 30, 'I': 15, 'J': 20, 'K': 20, 'L': 40, 'M': 40, 'N': 15,
        'O': 30, 'P': 40, 'Q': 30, 'R': 20, 'S': 20, 'T': 15, 'U': 15,
        'V': 15, 'W': 30, 'X': 15, 'Y': 15, 'Z': 25, 'AA': 25, 'AB': 25,
        'AC': 30, 'AD': 25
    }
    
    for col_letter, width in column_widths.items():
        ws_data.column_dimensions[col_letter].width = width
    
    # DROPDOWN OPTIONS
    
    # 1. IDENTIFICATION options (Column H)
    identification_options = [
        "African/Afrikan",
        "African-American/Black",
        "Afro-diasporic (Afro-Caribbean, Afro-Cubano, Black/African-American, Afro-Brazilian, etc.)",
        "Black/African-American",
        "Black/Afro-Diasporic",
        "Of African Descent (Afro-Caribbean, Afro-Cuban, Afro-Colombian, etc.)"
    ]
    
    # Write identification options to Options sheet
    ws_options['A1'] = 'IDENTIFICATION'
    ws_options['A1'].font = Font(bold=True)
    for i, option in enumerate(identification_options, 2):
        ws_options[f'A{i}'] = option
    
    # Create data validation for IDENTIFICATION (Column H)
    identification_dv = DataValidation(
        type="list",
        formula1=f"=Options!$A$2:$A${len(identification_options) + 1}",
        allow_blank=True
    )
    identification_dv.error = 'Please select a valid identification option'
    identification_dv.errorTitle = 'Invalid Identification'
    ws_data.add_data_validation(identification_dv)
    identification_dv.add(f"H2:H1000")  # Apply to column H
    
    # 2. GENDER options (Column I)
    gender_options = [
        "Female",
        "Male", 
        "Non-Binary",
        "Prefer not to say"
    ]
    
    ws_options['B1'] = 'GENDER'
    ws_options['B1'].font = Font(bold=True)
    for i, option in enumerate(gender_options, 2):
        ws_options[f'B{i}'] = option
    
    gender_dv = DataValidation(
        type="list",
        formula1=f"=Options!$B$2:$B${len(gender_options) + 1}",
        allow_blank=True
    )
    gender_dv.error = 'Please select a valid gender option'
    gender_dv.errorTitle = 'Invalid Gender'
    ws_data.add_data_validation(gender_dv)
    gender_dv.add(f"I2:I1000")  # Apply to column I
    
    # 3. PRIMARY INDUSTRY HOUSE options (Column L)
    primary_industry_options = [
        "☀️ Alternative Energy",
        "🌾 Agriculture/Sustainable Food Production / Land Management",
        "🏘 Community Development",
        "🛖 Eco-friendly Building",
        "💰 Alternative Economics",
        "🧑🏾‍🏫 Education & Cultural Preservation",
        "Environmental Justice/Advocacy",
        "♻️ Green Lifestyle",
        "🆘 Survival/Preparedness",
        "🗑 Waste",
        "💧Water",
        "🧘🏿‍♀️ Wholistic Health"
    ]
    
    ws_options['C1'] = 'PRIMARY INDUSTRY HOUSE'
    ws_options['C1'].font = Font(bold=True)
    for i, option in enumerate(primary_industry_options, 2):
        ws_options[f'C{i}'] = option
    
    primary_industry_dv = DataValidation(
        type="list",
        formula1=f"=Options!$C$2:$C${len(primary_industry_options) + 1}",
        allow_blank=True
    )
    primary_industry_dv.error = 'Please select a valid industry option'
    primary_industry_dv.errorTitle = 'Invalid Industry'
    ws_data.add_data_validation(primary_industry_dv)
    primary_industry_dv.add(f"L2:L1000")  # Apply to column L
    
    # 4. ADDITIONAL FOCUS AREAS options (Column M) - Same as primary industry
    additional_focus_dv = DataValidation(
        type="list",
        formula1=f"=Options!$C$2:$C${len(primary_industry_options) + 1}",
        allow_blank=True
    )
    additional_focus_dv.error = 'Please select valid focus areas (separate multiple with commas)'
    additional_focus_dv.errorTitle = 'Invalid Focus Areas'
    ws_data.add_data_validation(additional_focus_dv)
    additional_focus_dv.add(f"M2:M1000")  # Apply to column M
    
    # 5. US States and Canadian Provinces (Column S)
    us_states = [
        "Alabama", "Alaska", "Arizona", "Arkansas", "California", "Colorado",
        "Connecticut", "Delaware", "Florida", "Georgia", "Hawaii", "Idaho",
        "Illinois", "Indiana", "Iowa", "Kansas", "Kentucky", "Louisiana",
        "Maine", "Maryland", "Massachusetts", "Michigan", "Minnesota",
        "Mississippi", "Missouri", "Montana", "Nebraska", "Nevada",
        "New Hampshire", "New Jersey", "New Mexico", "New York",
        "North Carolina", "North Dakota", "Ohio", "Oklahoma", "Oregon",
        "Pennsylvania", "Rhode Island", "South Carolina", "South Dakota",
        "Tennessee", "Texas", "Utah", "Vermont", "Virginia", "Washington",
        "West Virginia", "Wisconsin", "Wyoming", "Washington D.C.",
        "Puerto Rico", "U.S. Virgin Islands", "Guam"
    ]
    
    canadian_provinces = [
        "Alberta", "British Columbia", "Manitoba", "New Brunswick",
        "Newfoundland and Labrador", "Northwest Territories", "Nova Scotia",
        "Nunavut", "Ontario", "Prince Edward Island", "Quebec", "Saskatchewan",
        "Yukon"
    ]
    
    state_province_options = sorted(us_states + canadian_provinces)
    
    ws_options['D1'] = 'STATE/PROVINCE'
    ws_options['D1'].font = Font(bold=True)
    for i, option in enumerate(state_province_options, 2):
        ws_options[f'D{i}'] = option
    
    state_province_dv = DataValidation(
        type="list",
        formula1=f"=Options!$D$2:$D${len(state_province_options) + 1}",
        allow_blank=True
    )
    state_province_dv.error = 'Please select a valid state or province'
    state_province_dv.errorTitle = 'Invalid State/Province'
    ws_data.add_data_validation(state_province_dv)
    state_province_dv.add(f"S2:S1000")  # Apply to column S (State/Province)
    state_province_dv.add(f"T2:T1000")  # Apply to column T (State) as well
    
    # 6. MEMBER LEVEL options (Column Z)
    member_level_options = [
        "Member",
        "Core Member",
        "Impact Member",
        "Legacy Member",
        "Featured Member",
        "Free Member"
    ]
    
    ws_options['E1'] = 'MEMBER LEVEL'
    ws_options['E1'].font = Font(bold=True)
    for i, option in enumerate(member_level_options, 2):
        ws_options[f'E{i}'] = option
    
    member_level_dv = DataValidation(
        type="list",
        formula1=f"=Options!$E$2:$E${len(member_level_options) + 1}",
        allow_blank=True
    )
    member_level_dv.error = 'Please select a valid member level'
    member_level_dv.errorTitle = 'Invalid Member Level'
    ws_data.add_data_validation(member_level_dv)
    member_level_dv.add(f"Z2:Z1000")  # Apply to column Z
    
    # 7. Yes/No dropdowns for boolean fields
    yes_no_options = ["Yes", "No", "TRUE", "FALSE", "true", "false"]
    
    ws_options['F1'] = 'YES/NO'
    ws_options['F1'].font = Font(bold=True)
    for i, option in enumerate(yes_no_options, 2):
        ws_options[f'F{i}'] = option
    
    yes_no_dv = DataValidation(
        type="list",
        formula1=f"=Options!$F$2:$F${len(yes_no_options) + 1}",
        allow_blank=True
    )
    yes_no_dv.error = 'Please select Yes or No'
    yes_no_dv.errorTitle = 'Invalid Value'
    ws_data.add_data_validation(yes_no_dv)
    yes_no_dv.add(f"W2:W1000")  # Include me on Global BSN Map
    yes_no_dv.add(f"AA2:AA1000")  # Paying Member
    yes_no_dv.add(f"AB2:AB1000")  # Equity Member
    yes_no_dv.add(f"AD2:AD1000")  # Send Need Payment Email
    
    # Hide the Options sheet
    ws_options.sheet_state = 'hidden'
    
    # Add instructions in row 2 of data sheet
    ws_data['A2'] = 'INSTRUCTIONS: Fill in member information below. Use dropdowns where provided. For multiple ADDITIONAL FOCUS AREAS, separate with commas.'
    ws_data['A2'].font = Font(italic=True, color="666666")
    
    # Freeze the header row
    ws_data.freeze_panes = 'A3'
    
    # Save the workbook
    output_file = '/Users/jerrybony/Documents/GitHub/black-sustainability-inc/data/BSN_Batch_Upload_Template_v2.xlsx'
    wb.save(output_file)
    print(f"✅ Successfully created: {output_file}")
    print("\nDropdown fields configured:")
    print("  - Column H: IDENTIFICATION")
    print("  - Column I: GENDER")
    print("  - Column L: PRIMARY INDUSTRY HOUSE")
    print("  - Column M: ADDITIONAL FOCUS AREAS")
    print("  - Column S: State/Province")
    print("  - Column T: State")
    print("  - Column Z: MEMBER LEVEL")
    print("  - Columns W, AA, AB, AD: Yes/No fields")
    print("\nNote: The 'Options' sheet contains all dropdown values and is hidden.")

if __name__ == "__main__":
    create_batch_upload_template()
